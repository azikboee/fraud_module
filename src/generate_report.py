import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os

def generate_report(file_path, report_path="Reports/fraud_report.xlsx"):
    print("Загрузка данных...")
    data = pd.read_csv(file_path)
    print("Колонки:", data.columns)

    if "fraud_prediction" not in data.columns:
        print(" В файле нет колонки fraud_prediction. Сначала запусти detect_fraud().")
        return

    print("Сохраняем Excel...")
    frauds = data[data["fraud_prediction"] == 1]
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="Все транзакции", index=False)
        frauds.to_excel(writer, sheet_name="Фродовые транзакции", index=False)
    print("Excel сохранён.")

    print("Строим график...")
    plt.figure(figsize=(12, 5))
    plt.scatter(data["timestamp"], data["amount"], alpha=0.5, label="Обычные")
    plt.scatter(frauds["timestamp"], frauds["amount"], color="red", label="Фрод", marker="x")
    plt.xlabel("Время")
    plt.ylabel("Сумма транзакции")
    plt.title("Фродовые транзакции")
    plt.legend()
    chart_path = "fraud_chart.png"
    plt.savefig(chart_path)
    plt.close()
    print("График сохранён.")

    print("Вставляем график в Excel...")
    from openpyxl import load_workbook
    wb = load_workbook(report_path)
    ws = wb.create_sheet("Графики")
    img = Image(chart_path)
    ws.add_image(img, "A1")
    wb.save(report_path)
    os.remove(chart_path)
    print(f" График добавлен в {report_path}")